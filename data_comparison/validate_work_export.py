from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from utils.work98_generator import (
    WORK98_LINE_END_ROW,
    WORK98_LINE_START_ROW,
    build_plan_render_rows,
    chunk_render_rows,
    employee_display_for_job,
    ordered_plan_groups,
    resolved_job_times,
    resolved_scheduled_job_times,
)


@dataclass
class ComparisonIssue:
    severity: str
    sheet: str
    row: int | None
    field: str
    expected: Any
    actual: Any
    message: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "severity": self.severity,
            "sheet": self.sheet,
            "row": self.row,
            "field": self.field,
            "expected": self.expected,
            "actual": self.actual,
            "message": self.message,
        }


def latest_file(folder: Path, pattern: str) -> Path:
    matches = sorted(folder.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    if not matches:
        raise FileNotFoundError(f"No files found for pattern: {pattern}")
    return matches[0]


def load_raw_jobs(raw_path: Path) -> list[dict[str, Any]]:
    with raw_path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def workbook_work_sheets(workbook_path: Path):
    wb = load_workbook(workbook_path, data_only=False)
    return [ws for ws in wb.worksheets if ws.title.startswith("Work")]


def normalize_date(value: Any) -> str | None:
    if value is None:
        return None
    if hasattr(value, "date"):
        return value.date().isoformat()
    return str(value)


def normalize_time(value: Any) -> str | None:
    if value is None:
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def actual_row_payload(ws, row_number: int) -> dict[str, Any]:
    return {
        "date": ws.cell(row=row_number, column=2).value,
        "employee": ws.cell(row=row_number, column=3).value,
        "in": ws.cell(row=row_number, column=4).value,
        "out": ws.cell(row=row_number, column=5).value,
        "scheduled_date": ws.cell(row=row_number, column=12).value,
        "scheduled_in": ws.cell(row=row_number, column=13).value,
        "scheduled_out": ws.cell(row=row_number, column=14).value,
    }


def expected_row_payload(job: dict[str, Any]) -> dict[str, Any]:
    start_time, end_time = resolved_job_times(job, None)
    scheduled_start_time, scheduled_end_time = resolved_scheduled_job_times(job, None)
    return {
        "date": start_time.date().isoformat() if start_time else None,
        "employee": employee_display_for_job(job, {}),
        "in": start_time.strftime("%Y-%m-%d %H:%M:%S") if start_time else None,
        "out": end_time.strftime("%Y-%m-%d %H:%M:%S") if end_time else None,
        "scheduled_date": scheduled_start_time.date().isoformat() if scheduled_start_time else None,
        "scheduled_in": scheduled_start_time.strftime("%Y-%m-%d %H:%M:%S") if scheduled_start_time else None,
        "scheduled_out": scheduled_end_time.strftime("%Y-%m-%d %H:%M:%S") if scheduled_end_time else None,
    }


def is_blank_payload(payload: dict[str, Any]) -> bool:
    return all(value in (None, "") for value in payload.values())


def compare_workbook_to_raw(workbook_path: Path, raw_path: Path) -> dict[str, Any]:
    jobs = load_raw_jobs(raw_path)
    contractors: dict[str, dict[str, str]] = {}
    expected_chunks: list[list[dict[str, Any] | None]] = []
    for plan_jobs in ordered_plan_groups(jobs, contractors):
        render_rows = build_plan_render_rows(plan_jobs, contractors)
        expected_chunks.extend(chunk_render_rows(render_rows, WORK98_LINE_END_ROW - WORK98_LINE_START_ROW + 1))

    worksheets = workbook_work_sheets(workbook_path)
    issues: list[ComparisonIssue] = []
    summary: dict[str, Any] = {
        "workbook": str(workbook_path),
        "raw": str(raw_path),
        "expected_sheet_count": len(expected_chunks),
        "actual_sheet_count": len(worksheets),
        "sheet_summaries": [],
    }

    if len(expected_chunks) != len(worksheets):
        issues.append(
            ComparisonIssue(
                severity="error",
                sheet="*",
                row=None,
                field="sheet_count",
                expected=len(expected_chunks),
                actual=len(worksheets),
                message="Sheet count differs between workbook and expected render.",
            )
        )

    for sheet_index, expected_chunk in enumerate(expected_chunks):
        if sheet_index >= len(worksheets):
            break
        ws = worksheets[sheet_index]
        sheet_summary = {
            "sheet": ws.title,
            "expected_render_rows": len(expected_chunk),
            "non_blank_actual_rows": 0,
            "blank_separator_rows": 0,
        }
        for offset, expected_job in enumerate(expected_chunk):
            row_number = WORK98_LINE_START_ROW + offset
            actual = actual_row_payload(ws, row_number)
            if expected_job is None:
                sheet_summary["blank_separator_rows"] += 1
                if not is_blank_payload(actual):
                    issues.append(
                        ComparisonIssue(
                            severity="error",
                            sheet=ws.title,
                            row=row_number,
                            field="separator_row",
                            expected="blank row",
                            actual=actual,
                            message="Expected a blank separator row between employees.",
                        )
                    )
                continue

            sheet_summary["non_blank_actual_rows"] += 1
            expected = expected_row_payload(expected_job)
            comparable_actual = {
                "date": normalize_date(actual["date"]),
                "employee": actual["employee"],
                "in": normalize_time(actual["in"]),
                "out": normalize_time(actual["out"]),
                "scheduled_date": normalize_date(actual["scheduled_date"]),
                "scheduled_in": normalize_time(actual["scheduled_in"]),
                "scheduled_out": normalize_time(actual["scheduled_out"]),
            }
            for field, expected_value in expected.items():
                actual_value = comparable_actual[field]
                if actual_value != expected_value:
                    issues.append(
                        ComparisonIssue(
                            severity="error",
                            sheet=ws.title,
                            row=row_number,
                            field=field,
                            expected=expected_value,
                            actual=actual_value,
                            message=f"Mismatch on {field}.",
                        )
                    )

        trailing_non_blank_rows: list[int] = []
        for row_number in range(WORK98_LINE_START_ROW + len(expected_chunk), WORK98_LINE_END_ROW + 1):
            if not is_blank_payload(actual_row_payload(ws, row_number)):
                trailing_non_blank_rows.append(row_number)
        if trailing_non_blank_rows:
            issues.append(
                ComparisonIssue(
                    severity="warning",
                    sheet=ws.title,
                    row=trailing_non_blank_rows[0],
                    field="trailing_rows",
                    expected="blank rows after expected content",
                    actual=trailing_non_blank_rows[:10],
                    message="Found non-blank rows after the expected rendered rows.",
                )
            )
        summary["sheet_summaries"].append(sheet_summary)

    summary["issue_count"] = len(issues)
    summary["issues"] = [issue.to_dict() for issue in issues]
    summary["status"] = "pass" if not issues else "fail"
    return summary


def main() -> None:
    folder = Path(__file__).resolve().parent
    workbook_path = latest_file(folder, "work_export_*.xlsx")
    raw_path = latest_file(folder, "work_export_raw_*.csv")
    summary = compare_workbook_to_raw(workbook_path, raw_path)

    json_path = folder / "validation_report.json"
    md_path = folder / "validation_report.md"
    json_path.write_text(json.dumps(summary, indent=2, ensure_ascii=True), encoding="utf-8")

    lines = [
        f"Status: {summary['status']}",
        f"Workbook: {summary['workbook']}",
        f"Raw: {summary['raw']}",
        f"Expected sheets: {summary['expected_sheet_count']}",
        f"Actual sheets: {summary['actual_sheet_count']}",
        f"Issue count: {summary['issue_count']}",
        "",
        "Sheet summaries:",
    ]
    for sheet_summary in summary["sheet_summaries"]:
        lines.append(
            f"- {sheet_summary['sheet']}: expected_render_rows={sheet_summary['expected_render_rows']}, "
            f"non_blank_actual_rows={sheet_summary['non_blank_actual_rows']}, "
            f"blank_separator_rows={sheet_summary['blank_separator_rows']}"
        )
    if summary["issues"]:
        lines.append("")
        lines.append("Issues:")
        for issue in summary["issues"][:200]:
            lines.append(
                f"- [{issue['severity']}] {issue['sheet']} row {issue['row']} field {issue['field']}: "
                f"expected={issue['expected']} actual={issue['actual']} | {issue['message']}"
            )
        if len(summary["issues"]) > 200:
            lines.append(f"- ... {len(summary['issues']) - 200} more issues omitted")
    else:
        lines.append("")
        lines.append("No issues found.")
    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json_path)
    print(md_path)
    print(summary['status'])
    print(summary['issue_count'])


if __name__ == "__main__":
    main()
