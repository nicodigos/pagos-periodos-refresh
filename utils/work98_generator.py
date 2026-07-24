import csv
import io
from collections import Counter
from collections import defaultdict
from copy import copy
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import pymysql
from openpyxl import load_workbook

from utils.ms_graph_excel import download_sharepoint_file_bytes, resolve_drive_id
from utils.pagos_periodos_sync import mysql_connection, parse_pagos_periodos_workbook
from utils.settings import get_bool_setting, get_setting

WORK98_TEMPLATE_SHEET_NAME = "Work98"
DATA_TEMPLATE_SHEET_NAME = "Data"
PAGOS_PERIODOS_BALANCE_SHEET_NAME = "Pagos Periodos"
TICKETS_BALANCE_SHEET_NAME = "Tickets"
BANK_PERIODICS_BALANCE_SHEET_NAME = "Bank Periodics"
WORK98_OUTPUT_PREFIX = "Work"
WORK98_LINE_START_ROW = 6
WORK98_LINE_END_ROW = 205
WORK98_LINE_CAPACITY = WORK98_LINE_END_ROW - WORK98_LINE_START_ROW + 1
DATA_HEADERS = (
    "Work",
    "Name Employee & vendor company",
    "Building",
    "Building & vendor company",
    "Date",
    "name employee",
    "Province",
    "total hours worked (number)",
    "hourly rate",
    "Total to pay",
    "Type of paiement",
    "Employee number",
    "Category",
    "Type of work",
    "Vendor Company",
)
BALANCE_HEADERS = (
    "Type of paiement",
    "Employee number",
    "Name employee",
    "Vendor Company",
    "Category",
    "Rows",
    "Total hours worked",
    "Total to pay",
)
BALANCE_SHEET_NAMES = (
    PAGOS_PERIODOS_BALANCE_SHEET_NAME,
    TICKETS_BALANCE_SHEET_NAME,
    BANK_PERIODICS_BALANCE_SHEET_NAME,
)
DEFAULT_SOURCE_TIMEZONE = "UTC"
DEFAULT_BUILDING_TIMEZONE = "America/Toronto"
BUILDING_TIMEZONE_BY_COUNTRY = {
    "canada": DEFAULT_BUILDING_TIMEZONE,
}
BUILDING_TIMEZONE_BY_PROVINCE = {
    ("canada", "alberta"): "America/Edmonton",
    ("canada", "british columbia"): "America/Vancouver",
    ("canada", "manitoba"): "America/Winnipeg",
    ("canada", "new brunswick"): "America/Moncton",
    ("canada", "newfoundland and labrador"): "America/St_Johns",
    ("canada", "newfoundland"): "America/St_Johns",
    ("canada", "nova scotia"): "America/Halifax",
    ("canada", "ontario"): "America/Toronto",
    ("canada", "prince edward island"): "America/Halifax",
    ("canada", "quebec"): "America/Montreal",
    ("canada", "saskatchewan"): "America/Regina",
    ("canada", "yukon"): "America/Whitehorse",
}
BUILDING_TIMEZONE_BY_CITY = {
    ("canada", "nunavut", "cambridge bay"): "America/Cambridge_Bay",
    ("canada", "nunavut", "iqaluit"): "America/Iqaluit",
    ("canada", "nunavut", "rankin inlet"): "America/Rankin_Inlet",
    ("canada", "ontario", "atikokan"): "America/Atikokan",
    ("canada", "ontario", "dryden"): "America/Winnipeg",
    ("canada", "ontario", "fort frances"): "America/Winnipeg",
    ("canada", "ontario", "kenora"): "America/Winnipeg",
    ("canada", "quebec", "blanc-sablon"): "America/Blanc-Sablon",
}
JOB_ACTUAL_START_TIME_KEYS = (
    "job_manual_start",
    "job_manual_start_time",
    "manual_start",
    "manual_start_time",
    "job_trim_start",
    "job_trimmed_start_time",
    "job_trimmed_start",
    "trimmed_start_time",
    "trimmed_start",
    "job_real_start_time",
    "job_real_start",
    "real_start_time",
    "real_start",
    "job_start_time",
    "start_time",
)
JOB_ACTUAL_END_TIME_KEYS = (
    "job_manual_end",
    "job_manual_end_time",
    "manual_end",
    "manual_end_time",
    "job_trim_end",
    "job_trimmed_end_time",
    "job_trimmed_end",
    "trimmed_end_time",
    "trimmed_end",
    "job_real_end_time",
    "job_real_end",
    "real_end_time",
    "real_end",
    "job_end_time",
    "end_time",
)
JOB_SCHEDULED_START_TIME_KEYS = (
    "job_scheduled_start_time",
    "job_scheduled_start",
    "scheduled_start_time",
    "scheduled_start",
)
JOB_SCHEDULED_END_TIME_KEYS = (
    "job_scheduled_end_time",
    "job_scheduled_end",
    "scheduled_end_time",
    "scheduled_end",
)
JOB_MANUAL_START_TIME_KEYS = (
    "job_manual_start_time",
    "job_manual_start",
    "manual_start_time",
    "manual_start",
)
JOB_MANUAL_END_TIME_KEYS = (
    "job_manual_end_time",
    "job_manual_end",
    "manual_end_time",
    "manual_end",
)
JOB_TRIMMED_START_TIME_KEYS = (
    "job_trim_start",
    "job_trimmed_start_time",
    "job_trimmed_start",
    "trimmed_start_time",
    "trimmed_start",
)
JOB_TRIMMED_END_TIME_KEYS = (
    "job_trim_end",
    "job_trimmed_end_time",
    "job_trimmed_end",
    "trimmed_end_time",
    "trimmed_end",
)
JOB_REAL_START_TIME_KEYS = (
    "job_real_start_time",
    "job_real_start",
    "real_start_time",
    "real_start",
    "job_start_time",
    "start_time",
)
JOB_REAL_END_TIME_KEYS = (
    "job_real_end_time",
    "job_real_end",
    "real_end_time",
    "real_end",
    "job_end_time",
    "end_time",
)
SCHEDULE_DATE_COLUMN = 12
SCHEDULE_START_COLUMN = 13
SCHEDULE_END_COLUMN = 14
SCHEDULE_HOURS_COLUMN = 15
SCHEDULE_HOURS_DIFF_COLUMN = 16
ATTENDED_COLUMN = 17
LATE_COLUMN = 18
REVIEW_COLUMN = 19
EXPIRED_COLUMN = 20
TRIMMED_END_FLAG_COLUMN = 21
MANUAL_START_COLUMN = 22
MANUAL_END_COLUMN = 23
TRIMMED_START_COLUMN = 24
TRIMMED_END_COLUMN = 25
REAL_START_COLUMN = 26
REAL_END_COLUMN = 27
EMPLOYEE_DATA_NAME_COLUMN = 28
LAST_WORK98_COLUMN = EMPLOYEE_DATA_NAME_COLUMN
SCHEDULE_HEADER_ROW = 5


def normalize_lookup_key(value: str) -> str:
    return " ".join((value or "").strip().lower().split())


def env_timezone(name: str, fallback: str) -> ZoneInfo:
    configured = str(get_setting(name, "")).strip() or fallback
    try:
        return ZoneInfo(configured)
    except ZoneInfoNotFoundError:
        return ZoneInfo(fallback)


def display_flag(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        normalized = value.strip().lower()
        if not normalized:
            return ""
        if normalized in {"1", "true", "yes", "y", "si", "s?"}:
            return "Yes"
        if normalized in {"0", "false", "no", "n"}:
            return "No"
    return "Yes" if bool(value) else "No"


def local_work98_template_path() -> Path:
    configured = str(get_setting("WORK98_TEMPLATE_LOCAL_PATH", "")).strip()
    if configured:
        return Path(configured)
    return Path(__file__).resolve().parent.parent / "sample.xlsx"


def work98_template_filename() -> str:
    configured = str(get_setting("WORK98_TEMPLATE_FILENAME", "")).strip()
    if configured:
        return configured
    return local_work98_template_path().name


def work98_template_bytes(token: str | None = None) -> bytes:
    sharepoint_path = str(get_setting("WORK98_TEMPLATE_PATH", "")).strip().strip("/")
    if sharepoint_path:
        if not token:
            raise RuntimeError("Microsoft token required to download WORK98_TEMPLATE_PATH")
        drive_id = resolve_drive_id(token)
        return download_sharepoint_file_bytes(sharepoint_path, token, drive_id=drive_id)

    local_path = local_work98_template_path()
    if not local_path.exists():
        raise RuntimeError(
            f"Work98 template not found: {local_path}. "
            "Add the template file to the deployment or configure WORK98_TEMPLATE_PATH in secrets."
        )
    return local_path.read_bytes()


def parse_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    value = str(value).strip()
    if not value or value.upper() == "NULL":
        return None
    return datetime.fromisoformat(value)


def first_datetime(job: dict[str, Any], keys: tuple[str, ...]) -> datetime | None:
    for key in keys:
        parsed = parse_datetime(job.get(key))
        if parsed is not None:
            return parsed
    return None


def building_timezone_name(job: dict[str, Any], building: dict[str, str] | None) -> str:
    for key in ("building_timezone", "timezone", "time_zone", "tz"):
        for source in (job, building or {}):
            value = str(source.get(key) or "").strip()
            if not value:
                continue
            try:
                ZoneInfo(value)
                return value
            except ZoneInfoNotFoundError:
                continue

    country = normalize_lookup_key(str(job.get("building_country") or (building or {}).get("country", "")))
    province = normalize_lookup_key(str(job.get("building_province") or (building or {}).get("province", "")))
    city = normalize_lookup_key(str(job.get("building_city") or (building or {}).get("city", "")))

    return (
        BUILDING_TIMEZONE_BY_CITY.get((country, province, city))
        or BUILDING_TIMEZONE_BY_PROVINCE.get((country, province))
        or BUILDING_TIMEZONE_BY_COUNTRY.get(country)
        or str(get_setting("DEFAULT_BUILDING_TIMEZONE", "")).strip()
        or DEFAULT_BUILDING_TIMEZONE
    )


def resolve_building_timezone(job: dict[str, Any], building: dict[str, str] | None) -> ZoneInfo:
    try:
        return ZoneInfo(building_timezone_name(job, building))
    except ZoneInfoNotFoundError:
        return ZoneInfo(DEFAULT_BUILDING_TIMEZONE)


def convert_job_datetime_to_building_time(value: datetime | None, building_timezone: ZoneInfo) -> datetime | None:
    if value is None:
        return None

    if value.tzinfo is None:
        value = value.replace(tzinfo=env_timezone("JOB_SOURCE_TIMEZONE", DEFAULT_SOURCE_TIMEZONE))

    return value.astimezone(building_timezone).replace(tzinfo=None)


def resolved_job_times(job: dict[str, Any], building: dict[str, str] | None) -> tuple[datetime | None, datetime | None]:
    building_timezone = resolve_building_timezone(job, building)
    start_time = convert_job_datetime_to_building_time(first_datetime(job, JOB_ACTUAL_START_TIME_KEYS), building_timezone)
    end_time = convert_job_datetime_to_building_time(first_datetime(job, JOB_ACTUAL_END_TIME_KEYS), building_timezone)
    return start_time, end_time


def resolved_scheduled_job_times(
    job: dict[str, Any],
    building: dict[str, str] | None,
) -> tuple[datetime | None, datetime | None]:
    building_timezone = resolve_building_timezone(job, building)
    start_time = convert_job_datetime_to_building_time(first_datetime(job, JOB_SCHEDULED_START_TIME_KEYS), building_timezone)
    end_time = convert_job_datetime_to_building_time(first_datetime(job, JOB_SCHEDULED_END_TIME_KEYS), building_timezone)
    return start_time, end_time


def resolved_job_source_times(
    job: dict[str, Any],
    building: dict[str, str] | None,
    start_keys: tuple[str, ...],
    end_keys: tuple[str, ...],
) -> tuple[datetime | None, datetime | None]:
    building_timezone = resolve_building_timezone(job, building)
    start_time = convert_job_datetime_to_building_time(first_datetime(job, start_keys), building_timezone)
    end_time = convert_job_datetime_to_building_time(first_datetime(job, end_keys), building_timezone)
    return start_time, end_time


def contractor_display_name(job: dict[str, Any], contractor: dict[str, str] | None) -> str:
    for key in ("assigned_user_pagos_periodos_id",):
        value = (job.get(key) or "").strip()
        if value:
            return value
    full_name = " ".join(
        part
        for part in [job.get("worker_first_name", "").strip(), job.get("worker_last_name", "").strip()]
        if part
    )
    if full_name:
        return full_name
    if contractor:
        value = (contractor.get("name") or "").strip()
        if value:
            return value
    return (job.get("worker_username") or "").strip()


def building_display_name(job: dict[str, Any], building: dict[str, str] | None) -> str:
    for key in ("building_pagos_periodos_id", "building_address"):
        value = (job.get(key) or "").strip()
        if value:
            return value
    if building:
        return (building.get("name") or "").strip()
    return ""


def app_employee_name(job: dict[str, Any]) -> str:
    full_name = " ".join(
        part
        for part in [str(job.get("worker_first_name") or "").strip(), str(job.get("worker_last_name") or "").strip()]
        if part
    )
    if full_name:
        return full_name
    return str(job.get("worker_username") or "").strip()


def build_lookup_dict(records: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    for record in records:
        key = normalize_lookup_key(record.get("name", ""))
        if key:
            out[key] = record
    return out


def load_sample_work98_inputs() -> tuple[list[dict[str, str]], dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    sample_csv_path = Path(__file__).resolve().parent.parent / "sample_db.csv"
    if not sample_csv_path.exists():
        raise RuntimeError(f"Sample DB CSV not found: {sample_csv_path}")

    with sample_csv_path.open(newline="", encoding="utf-8") as handle:
        jobs = list(csv.DictReader(handle))

    workbook_data = parse_pagos_periodos_workbook(work98_template_bytes())
    return jobs, build_lookup_dict(workbook_data["contractors"]), build_lookup_dict(workbook_data["buildings"])


def datetime_for_match(value: Any) -> str:
    parsed = parse_datetime(value)
    if parsed is None:
        return ""
    return parsed.strftime("%Y-%m-%d %H:%M:%S")


def job_schedule_match_key(job: dict[str, Any]) -> tuple[str, str, str, str, str]:
    return (
        str(job.get("plan_user_id") or job.get("schedule_plan_user_id") or "").strip(),
        str(job.get("plan_id") or job.get("schedule_plan_id") or "").strip(),
        str(job.get("job_route_id") or job.get("schedule_route_id") or "").strip(),
        datetime_for_match(
            job.get("job_scheduled_start")
            or job.get("job_scheduled_start_time")
            or job.get("schedule_start_at")
        ),
        datetime_for_match(
            job.get("job_scheduled_end")
            or job.get("job_scheduled_end_time")
            or job.get("schedule_end_at")
        ),
    )


def normalize_job_record(job: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(job)
    if not normalized.get("assigned_user_pagos_periodos_id"):
        normalized["assigned_user_pagos_periodos_id"] = (
            normalized.get("job_pagos_periodos_id")
            or normalized.get("route_assignment_pagos_periodos_id")
            or ""
        )
    return normalized


def normalize_job_records(jobs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [normalize_job_record(job) for job in jobs]


def view_includes_schedule_only_rows(jobs: list[dict[str, Any]]) -> bool:
    for job in jobs:
        if job.get("job_id"):
            continue
        if (
            job.get("job_scheduled_start")
            or job.get("job_scheduled_start_time")
            or job.get("job_scheduled_end")
            or job.get("job_scheduled_end_time")
        ):
            return True
    return False


def resolved_filter_datetimes(
    job: dict[str, Any],
    building: dict[str, str] | None,
) -> tuple[datetime | None, datetime | None]:
    building_timezone = resolve_building_timezone(job, building)
    scheduled_start = convert_job_datetime_to_building_time(first_datetime(job, JOB_SCHEDULED_START_TIME_KEYS), building_timezone)
    scheduled_end = convert_job_datetime_to_building_time(first_datetime(job, JOB_SCHEDULED_END_TIME_KEYS), building_timezone)
    fallback_start = convert_job_datetime_to_building_time(first_datetime(job, JOB_ACTUAL_START_TIME_KEYS), building_timezone)
    fallback_end = convert_job_datetime_to_building_time(first_datetime(job, JOB_ACTUAL_END_TIME_KEYS), building_timezone)
    return scheduled_start or fallback_start, scheduled_end or fallback_end


def enrich_jobs_for_filtering(
    jobs: list[dict[str, Any]],
    buildings: dict[str, dict[str, str]],
) -> list[dict[str, Any]]:
    enriched: list[dict[str, Any]] = []
    for job in jobs:
        normalized = dict(job)
        building = find_building_lookup(normalized, buildings)
        filter_start, filter_end = resolved_filter_datetimes(normalized, building)
        normalized["filter_start_datetime"] = filter_start.isoformat(sep=" ") if filter_start else ""
        normalized["filter_end_datetime"] = filter_end.isoformat(sep=" ") if filter_end else ""
        normalized["filter_start_date"] = filter_start.date().isoformat() if filter_start else ""
        normalized["filter_end_date"] = filter_end.date().isoformat() if filter_end else ""
        enriched.append(normalized)
    return enriched


def filter_jobs_by_date_range(
    jobs: list[dict[str, Any]],
    start_date: date | None = None,
    end_date: date | None = None,
) -> list[dict[str, Any]]:
    if start_date is None and end_date is None:
        return jobs

    filtered: list[dict[str, Any]] = []
    for job in jobs:
        filter_start = parse_datetime(job.get("filter_start_datetime"))
        filter_end = parse_datetime(job.get("filter_end_datetime"))
        if filter_start is None and filter_end is None:
            continue
        effective_start = (filter_start or filter_end).date()
        effective_end = (filter_end or filter_start).date()
        if start_date is not None and effective_end < start_date:
            continue
        if end_date is not None and effective_start > end_date:
            continue
        filtered.append(job)
    return filtered


def fetch_past_schedules_from_db(cursor: pymysql.cursors.DictCursor) -> list[dict[str, Any]]:
    sql = """
        SELECT
            CONCAT('schedule-only:', s.id) AS job_id,
            BIN_TO_UUID(s.plan_user_id) AS plan_user_id,
            BIN_TO_UUID(s.plan_user_id) AS job_plan_user_id,
            BIN_TO_UUID(s.route_id) AS job_route_id,
            BIN_TO_UUID(s.plan_id) AS plan_id,
            NULL AS job_route_assignment_version_id,
            NULL AS job_start_time,
            NULL AS job_end_time,
            s.scheduled_start_at AS job_scheduled_start,
            s.scheduled_end_at AS job_scheduled_end,
            0 AS job_active,
            0 AS job_expired,
            0 AS job_needs_review,
            s.scheduled_start_at AS job_created_at,
            s.route_name AS job_route_name,
            s.route_description AS job_route_description,
            rav.schedule_label AS job_route_schedule_label,
            COALESCE(rav.pagos_periodos_id, '') AS assigned_user_pagos_periodos_id,
            COALESCE(rav.type_of_payment, '') AS job_type_of_payment,
            rav.hourly_rate AS job_hourly_rate,
            rav.scheduled_hours AS job_scheduled_hours,
            pu.user_id AS assigned_user_id,
            pu.plan_id AS assigned_plan_id,
            pu.assigned_at AS plan_user_assigned_at,
            pu.active AS plan_user_active,
            COALESCE(p.first_name, '') AS worker_first_name,
            COALESCE(p.last_name, '') AS worker_last_name,
            COALESCE(u.username, '') AS worker_username,
            pl.status AS plan_status,
            pl.start_date AS plan_start_date,
            pl.end_date AS plan_end_date,
            pl.description AS plan_description,
            pl.po_code AS plan_po_code,
            pl.budget AS plan_budget,
            pt.name AS plan_type_name,
            BIN_TO_UUID(b.id) AS building_id,
            COALESCE(b.public_building_id, '') AS building_number,
            COALESCE(b.public_expense_id, '') AS building_expense_id,
            b.address AS building_address,
            COALESCE(b.pagos_periodos_id, '') AS building_pagos_periodos_id,
            COALESCE(c.city_name, '') AS building_city,
            COALESCE(pr.province_name, '') AS building_province,
            COALESCE(co.country_name, '') AS building_country
        FROM scheduled_shift_instances s
        LEFT JOIN route_assignment_versions rav
            ON rav.route_id = s.route_id
           AND rav.plan_user_id = s.plan_user_id
           AND rav.current = 1
        LEFT JOIN plan_users pu
            ON pu.id = s.plan_user_id
        LEFT JOIN users u
            ON u.id = pu.user_id
        LEFT JOIN persons p
            ON p.id = u.person_id
        LEFT JOIN plans pl
            ON pl.id = s.plan_id
        LEFT JOIN plan_types pt
            ON pt.id = pl.type_id
        LEFT JOIN buildings b
            ON b.id = pl.building_id
        LEFT JOIN cities c
            ON c.id = b.city_id
        LEFT JOIN provinces pr
            ON pr.id = c.province_id
        LEFT JOIN countries co
            ON co.id = pr.country_id
        WHERE s.schedule_status <> 'cancelled'
          AND s.scheduled_end_at <= UTC_TIMESTAMP()
        ORDER BY s.scheduled_start_at DESC, s.worker_name ASC, s.route_name ASC
    """
    cursor.execute(sql)
    return list(cursor.fetchall())


def merge_jobs_with_schedules(
    jobs: list[dict[str, Any]],
    schedules: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    matched_jobs = Counter()
    for job in jobs:
        key = job_schedule_match_key(job)
        if key[3] and key[4]:
            matched_jobs[key] += 1

    merged_jobs = list(jobs)
    for schedule in schedules:
        key = job_schedule_match_key(schedule)
        if not key[3] or not key[4]:
            continue
        if matched_jobs.get(key, 0):
            matched_jobs[key] -= 1
            continue
        merged_jobs.append(schedule)
    return merged_jobs


def fetch_work98_inputs_from_db() -> tuple[list[dict[str, Any]], dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    with mysql_connection() as connection:
        with connection.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute(
                "SELECT * FROM vw_jobs_with_plan_building_and_pagos_periodos ORDER BY job_created_at DESC"
            )
            jobs = normalize_job_records(list(cursor.fetchall()))
            schedules: list[dict[str, Any]] = []
            if not view_includes_schedule_only_rows(jobs):
                schedules = normalize_job_records(fetch_past_schedules_from_db(cursor))
            cursor.execute(
                "SELECT name, hourly_rate, type_of_payment, employee_sub_grouping, category "
                "FROM pagos_periodos_contractors"
            )
            contractors = build_lookup_dict(list(cursor.fetchall()))
            cursor.execute(
                "SELECT name, invoicing, expense_id, building_number, client, vendor, province, budget, "
                "type_of_plan, po, icn FROM pagos_periodos_buildings"
            )
            buildings = build_lookup_dict(list(cursor.fetchall()))
    jobs = merge_jobs_with_schedules(jobs, schedules)
    jobs = enrich_jobs_for_filtering(jobs, buildings)
    return jobs, contractors, buildings


def load_work98_inputs() -> tuple[list[dict[str, Any]], dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    use_sample = get_bool_setting("WORK98_USE_SAMPLE_DATA", False)
    if use_sample:
        return load_sample_work98_inputs()
    return fetch_work98_inputs_from_db()


def find_contractor_lookup(job: dict[str, Any], contractors: dict[str, dict[str, str]]) -> dict[str, str] | None:
    candidate_keys = [
        job.get("assigned_user_pagos_periodos_id", ""),
        " ".join(
            part
            for part in [job.get("worker_first_name", "").strip(), job.get("worker_last_name", "").strip()]
            if part
        ),
        job.get("worker_username", ""),
    ]
    for candidate in candidate_keys:
        key = normalize_lookup_key(candidate)
        if key and key in contractors:
            return contractors[key]
    return None


def find_building_lookup(job: dict[str, Any], buildings: dict[str, dict[str, str]]) -> dict[str, str] | None:
    candidate_keys = [
        job.get("building_pagos_periodos_id", ""),
        job.get("building_address", ""),
    ]
    for candidate in candidate_keys:
        key = normalize_lookup_key(candidate)
        if key and key in buildings:
            return buildings[key]
    return None


def employee_display_for_job(job: dict[str, Any], contractors: dict[str, dict[str, str]]) -> str:
    cnetbms_name = app_employee_name(job)
    pagos_periodos_name = str(job.get("assigned_user_pagos_periodos_id") or "").strip()
    if cnetbms_name and pagos_periodos_name:
        return f"{cnetbms_name} ({pagos_periodos_name})"
    return cnetbms_name or pagos_periodos_name


def ordered_plan_groups(
    jobs: list[dict[str, Any]],
    contractors: dict[str, dict[str, str]],
) -> list[list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    sort_key_by_plan: dict[str, tuple[datetime | None, str]] = {}

    for job in jobs:
        plan_id = str(job.get("plan_id") or "").strip()
        if not plan_id:
            continue
        grouped[plan_id].append(job)
        job_created_at = parse_datetime(job.get("job_created_at"))
        existing = sort_key_by_plan.get(plan_id)
        candidate_key = (job_created_at, plan_id)
        if existing is None or (candidate_key[0] and existing[0] and candidate_key[0] > existing[0]) or (candidate_key[0] and existing[0] is None):
            sort_key_by_plan[plan_id] = candidate_key

    ordered_plan_ids = [
        plan_id
        for plan_id, _ in sorted(
            sort_key_by_plan.items(),
            key=lambda item: (item[1][0] or datetime.min),
            reverse=True,
        )
    ]

    ordered_groups: list[list[dict[str, Any]]] = []
    for plan_id in ordered_plan_ids:
        ordered_groups.append(
            sorted(
                grouped[plan_id],
                key=lambda job: (
                    normalize_lookup_key(employee_display_for_job(job, contractors)),
                    first_datetime(job, JOB_ACTUAL_START_TIME_KEYS)
                    or first_datetime(job, JOB_SCHEDULED_START_TIME_KEYS)
                    or datetime.min,
                    str(job.get("job_id") or ""),
                ),
            )
        )
    return ordered_groups


def set_cell_value(cell, value: Any) -> None:
    value = empty_to_none(value)
    if value is None:
        cell._value = None
        cell.data_type = "n"
        return
    cell.value = value


def clear_work98_line_rows(ws) -> None:
    for row in range(WORK98_LINE_START_ROW, WORK98_LINE_END_ROW + 1):
        for column in range(2, LAST_WORK98_COLUMN + 1):
            set_cell_value(ws.cell(row=row, column=column), None)


def populate_schedule_headers(ws) -> None:
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=SCHEDULE_DATE_COLUMN), "Scheduled Date")
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=SCHEDULE_START_COLUMN), "Scheduled In")
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=SCHEDULE_END_COLUMN), "Scheduled Out")
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=SCHEDULE_HOURS_COLUMN), "Scheduled Hours")
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=SCHEDULE_HOURS_DIFF_COLUMN), "Worked vs Scheduled")
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=ATTENDED_COLUMN), "Attended")
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=LATE_COLUMN), "Late")
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=REVIEW_COLUMN), "Review")
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=EXPIRED_COLUMN), "Expired")
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=TRIMMED_END_FLAG_COLUMN), "Trimmed end")
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=MANUAL_START_COLUMN), "Manual start")
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=MANUAL_END_COLUMN), "Manual end")
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=TRIMMED_START_COLUMN), "Trimmed start")
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=TRIMMED_END_COLUMN), "Trimmed end time")
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=REAL_START_COLUMN), "Real start")
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=REAL_END_COLUMN), "Real end")
    set_cell_value(ws.cell(row=SCHEDULE_HEADER_ROW, column=EMPLOYEE_DATA_NAME_COLUMN), "Employee data name")


def populate_work98_header(ws, jobs: list[dict[str, Any]], buildings: dict[str, dict[str, str]]) -> None:
    first_job = jobs[0]
    building = find_building_lookup(first_job, buildings)

    set_cell_value(ws["C2"], (building or {}).get("client", ""))
    set_cell_value(ws["E2"], (first_job.get("plan_type_name") or "").strip())
    set_cell_value(ws["H2"], (building or {}).get("po", ""))
    set_cell_value(ws["C3"], (first_job.get("plan_description") or "").strip())
    set_cell_value(ws["E3"], (building or {}).get("building_number", ""))
    set_cell_value(ws["K3"], (building or {}).get("vendor", ""))
    set_cell_value(ws["C4"], building_display_name(first_job, building))
    set_cell_value(ws["E4"], (building or {}).get("expense_id", ""))


def empty_to_none(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def float_or_blank(value: str) -> float | str | None:
    value = (value or "").strip()
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return value


def clone_cell_style(source_cell, target_cell) -> None:
    target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format


def sheet_title_for_data(sheet_title: str) -> str:
    if sheet_title.startswith(WORK98_OUTPUT_PREFIX):
        suffix = sheet_title[len(WORK98_OUTPUT_PREFIX):].strip()
        if suffix.isdigit():
            return f"{WORK98_OUTPUT_PREFIX} {suffix}"
    return sheet_title


def build_data_row(
    sheet_title: str,
    job: dict[str, Any],
    contractor: dict[str, str] | None,
    building: dict[str, str] | None,
) -> list[Any]:
    employee_name = contractor_display_name(job, contractor)
    vendor_company = (building or {}).get("vendor", "")
    building_name = building_display_name(job, building)
    start_time, end_time = resolved_job_times(job, building)
    hours_worked = None
    if start_time and end_time:
        hours_worked = round((end_time - start_time).total_seconds() / 3600, 4)

    hourly_rate = float_or_blank((contractor or {}).get("hourly_rate", ""))
    total_to_pay = None
    if isinstance(hours_worked, (int, float)) and isinstance(hourly_rate, (int, float)):
        total_to_pay = round(hours_worked * hourly_rate, 2)

    employee_number = (contractor or {}).get("employee_sub_grouping", "")
    type_of_work = (job.get("plan_type_name") or "").strip()
    if not type_of_work and building:
        type_of_work = (building.get("type_of_plan") or "").strip()

    return [
        sheet_title_for_data(sheet_title),
        " ".join(part for part in [employee_name, vendor_company] if part).strip(),
        building_name,
        " ".join(part for part in [building_name, vendor_company] if part).strip(),
        start_time.date() if start_time else None,
        employee_name,
        (building or {}).get("province", "") or (job.get("building_province") or ""),
        hours_worked,
        hourly_rate,
        total_to_pay,
        (contractor or {}).get("type_of_payment", ""),
        employee_number,
        (contractor or {}).get("category", ""),
        type_of_work,
        vendor_company,
    ]


def populate_work98_rows(
    ws,
    jobs: list[dict[str, Any] | None],
    contractors: dict[str, dict[str, str]],
    buildings: dict[str, dict[str, str]],
) -> None:
    clear_work98_line_rows(ws)
    populate_schedule_headers(ws)

    for offset, job in enumerate(jobs):
        row_number = WORK98_LINE_START_ROW + offset
        if job is None:
            continue
        contractor = find_contractor_lookup(job, contractors)
        building = find_building_lookup(job, buildings)
        start_time, end_time = resolved_job_times(job, building)
        scheduled_start_time, scheduled_end_time = resolved_scheduled_job_times(job, building)
        manual_start_time, manual_end_time = resolved_job_source_times(job, building, JOB_MANUAL_START_TIME_KEYS, JOB_MANUAL_END_TIME_KEYS)
        trimmed_start_time, trimmed_end_time = resolved_job_source_times(job, building, JOB_TRIMMED_START_TIME_KEYS, JOB_TRIMMED_END_TIME_KEYS)
        real_start_time, real_end_time = resolved_job_source_times(job, building, JOB_REAL_START_TIME_KEYS, JOB_REAL_END_TIME_KEYS)

        display_date = start_time.date() if start_time else (scheduled_start_time.date() if scheduled_start_time else None)
        set_cell_value(ws.cell(row=row_number, column=2), display_date)
        set_cell_value(ws.cell(row=row_number, column=3), employee_display_for_job(job, contractors))
        set_cell_value(ws.cell(row=row_number, column=4), start_time if start_time else None)
        set_cell_value(ws.cell(row=row_number, column=5), end_time if end_time else None)
        ws.cell(row=row_number, column=6).value = f'=IF(OR(D{row_number}="",E{row_number}=""),"",E{row_number}-D{row_number})'
        ws.cell(row=row_number, column=7).value = (
            f'=IF(F{row_number}="","",IF(((F{row_number}-INT(F{row_number}))*24>=6),'
            f'(((F{row_number}-INT(F{row_number}))*24)-0.5),((F{row_number}-INT(F{row_number}))*24)))'
        )
        set_cell_value(ws.cell(row=row_number, column=8), float_or_blank((contractor or {}).get("hourly_rate", "")))
        ws.cell(row=row_number, column=9).value = f'=IF(OR(G{row_number}="",H{row_number}=""),"",G{row_number}*H{row_number})'
        set_cell_value(ws.cell(row=row_number, column=10), (contractor or {}).get("category", ""))
        set_cell_value(
            ws.cell(row=row_number, column=SCHEDULE_DATE_COLUMN),
            scheduled_start_time.date() if scheduled_start_time else None,
        )
        set_cell_value(ws.cell(row=row_number, column=SCHEDULE_START_COLUMN), scheduled_start_time if scheduled_start_time else None)
        set_cell_value(ws.cell(row=row_number, column=SCHEDULE_END_COLUMN), scheduled_end_time if scheduled_end_time else None)
        ws.cell(row=row_number, column=SCHEDULE_HOURS_COLUMN).value = (
            f'=IF(OR(M{row_number}="",N{row_number}=""),"",IF(((N{row_number}-M{row_number})*24>=6),((N{row_number}-M{row_number})*24)-0.5,((N{row_number}-M{row_number})*24)))'
        )
        ws.cell(row=row_number, column=SCHEDULE_HOURS_DIFF_COLUMN).value = (
            f'=IF(OR(G{row_number}="",O{row_number}=""),"",G{row_number}-O{row_number})'
        )
        ws.cell(row=row_number, column=ATTENDED_COLUMN).value = (
            f'=IF(OR(M{row_number}<>"",N{row_number}<>""),IF(OR(D{row_number}<>"",E{row_number}<>""),"Yes","No"),"")'
        )
        ws.cell(row=row_number, column=LATE_COLUMN).value = (
            f'=IF(OR(Q{row_number}="",Q{row_number}="No",M{row_number}="",D{row_number}=""),"",IF(D{row_number}>M{row_number},"Yes","No"))'
        )
        set_cell_value(ws.cell(row=row_number, column=REVIEW_COLUMN), display_flag(job.get("job_needs_review")))
        set_cell_value(ws.cell(row=row_number, column=EXPIRED_COLUMN), display_flag(job.get("job_expired")))
        set_cell_value(ws.cell(row=row_number, column=TRIMMED_END_FLAG_COLUMN), display_flag(trimmed_end_time))
        set_cell_value(ws.cell(row=row_number, column=MANUAL_START_COLUMN), manual_start_time)
        set_cell_value(ws.cell(row=row_number, column=MANUAL_END_COLUMN), manual_end_time)
        set_cell_value(ws.cell(row=row_number, column=TRIMMED_START_COLUMN), trimmed_start_time)
        set_cell_value(ws.cell(row=row_number, column=TRIMMED_END_COLUMN), trimmed_end_time)
        set_cell_value(ws.cell(row=row_number, column=REAL_START_COLUMN), real_start_time)
        set_cell_value(ws.cell(row=row_number, column=REAL_END_COLUMN), real_end_time)
        set_cell_value(ws.cell(row=row_number, column=EMPLOYEE_DATA_NAME_COLUMN), app_employee_name(job))

        ws.cell(row=row_number, column=2).number_format = "yyyy-mm-dd"
        ws.cell(row=row_number, column=4).number_format = "hh:mm:ss"
        ws.cell(row=row_number, column=5).number_format = "hh:mm:ss"
        ws.cell(row=row_number, column=6).number_format = "[h]:mm:ss"
        ws.cell(row=row_number, column=SCHEDULE_DATE_COLUMN).number_format = "yyyy-mm-dd"
        ws.cell(row=row_number, column=SCHEDULE_START_COLUMN).number_format = "hh:mm:ss"
        ws.cell(row=row_number, column=SCHEDULE_END_COLUMN).number_format = "hh:mm:ss"
        ws.cell(row=row_number, column=SCHEDULE_HOURS_COLUMN).number_format = "0.00"
        ws.cell(row=row_number, column=SCHEDULE_HOURS_DIFF_COLUMN).number_format = "0.00"
        ws.cell(row=row_number, column=MANUAL_START_COLUMN).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row=row_number, column=MANUAL_END_COLUMN).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row=row_number, column=TRIMMED_START_COLUMN).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row=row_number, column=TRIMMED_END_COLUMN).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row=row_number, column=REAL_START_COLUMN).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row=row_number, column=REAL_END_COLUMN).number_format = "yyyy-mm-dd hh:mm:ss"


def build_plan_render_rows(
    jobs: list[dict[str, Any]],
    contractors: dict[str, dict[str, str]],
) -> list[dict[str, Any] | None]:
    rows: list[dict[str, Any] | None] = []
    previous_employee_key = None
    for job in jobs:
        employee_key = normalize_lookup_key(employee_display_for_job(job, contractors))
        if rows and previous_employee_key is not None and employee_key != previous_employee_key:
            rows.append(None)
        rows.append(job)
        previous_employee_key = employee_key
    return rows


def chunk_render_rows(values: list[dict[str, Any] | None], size: int) -> list[list[dict[str, Any] | None]]:
    chunks: list[list[dict[str, Any] | None]] = []
    current_chunk: list[dict[str, Any] | None] = []
    for value in values:
        if len(current_chunk) >= size:
            chunks.append(current_chunk)
            current_chunk = []
        if value is None and not current_chunk:
            continue
        current_chunk.append(value)
    if current_chunk:
        chunks.append(current_chunk)
    return chunks or [[]]


def build_dict_rows_csv_bytes(rows: list[dict[str, Any]]) -> bytes:
    headers: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in headers:
                headers.append(key)

    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(buffer, fieldnames=headers)
    writer.writeheader()
    for row in rows:
        writer.writerow({header: row.get(header) for header in headers})
    return buffer.getvalue().encode("utf-8-sig")


def build_raw_jobs_csv_bytes(jobs: list[dict[str, Any]]) -> bytes:
    return build_dict_rows_csv_bytes(jobs)


def ensure_data_sheet(workbook):
    if DATA_TEMPLATE_SHEET_NAME in workbook.sheetnames:
        return workbook[DATA_TEMPLATE_SHEET_NAME]

    ws = workbook.create_sheet(DATA_TEMPLATE_SHEET_NAME)
    for col_index, header in enumerate(DATA_HEADERS, start=1):
        ws.cell(row=1, column=col_index).value = header
    return ws


def populate_data_sheet(ws, data_rows: list[list[Any]]) -> None:
    for col_index, header in enumerate(DATA_HEADERS, start=1):
        if ws.cell(row=1, column=col_index).value in (None, ""):
            ws.cell(row=1, column=col_index).value = header

    template_cells = [ws.cell(row=2, column=col_index) for col_index in range(1, len(DATA_HEADERS) + 1)]
    template_row_height = ws.row_dimensions[2].height

    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)

    if len(data_rows) > 1:
        ws.insert_rows(3, len(data_rows) - 1)

    row_count = max(len(data_rows), 1)
    for row_offset in range(2, row_count + 2):
        values = data_rows[row_offset - 2] if row_offset - 2 < len(data_rows) else [None] * len(DATA_HEADERS)
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_offset, column=col_index)
            clone_cell_style(template_cells[col_index - 1], cell)
            set_cell_value(cell, value)
        if template_row_height is not None:
            ws.row_dimensions[row_offset].height = template_row_height
        ws.cell(row=row_offset, column=5).number_format = "yyyy-mm-dd"
        ws.cell(row=row_offset, column=8).number_format = "0.0000"
        ws.cell(row=row_offset, column=9).number_format = "0.00"
        ws.cell(row=row_offset, column=10).number_format = "0.00"


def balance_category(data_row: list[Any]) -> str:
    row = dict(zip(DATA_HEADERS, data_row))
    combined = normalize_lookup_key(
        " ".join(
            str(row.get(header) or "")
            for header in (
                "Building",
                "Building & vendor company",
                "Type of work",
                "Vendor Company",
            )
        )
    )
    type_of_work = normalize_lookup_key(str(row.get("Type of work") or ""))

    if "bank periodic" in combined or "banks periodic" in combined or ("bank" in combined and "periodic" in combined):
        return BANK_PERIODICS_BALANCE_SHEET_NAME
    if any(term in type_of_work for term in ("one shot", "oneshot", "ticket", "service call", "project work", "project")):
        return TICKETS_BALANCE_SHEET_NAME
    return PAGOS_PERIODOS_BALANCE_SHEET_NAME


def build_balance_rows(data_rows: list[list[Any]], category: str) -> list[list[Any]]:
    grouped: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    for data_row in data_rows:
        if balance_category(data_row) != category:
            continue

        row = dict(zip(DATA_HEADERS, data_row))
        key = (
            str(row.get("Type of paiement") or "").strip(),
            str(row.get("Employee number") or "").strip(),
            str(row.get("name employee") or "").strip(),
            str(row.get("Vendor Company") or "").strip(),
            str(row.get("Category") or "").strip(),
        )
        summary = grouped.setdefault(
            key,
            {
                "rows": 0,
                "hours": 0.0,
                "total": 0.0,
            },
        )
        summary["rows"] += 1
        hours = row.get("total hours worked (number)")
        total = row.get("Total to pay")
        if isinstance(hours, (int, float)):
            summary["hours"] += hours
        if isinstance(total, (int, float)):
            summary["total"] += total

    out: list[list[Any]] = []
    for key, summary in sorted(grouped.items(), key=lambda item: tuple(normalize_lookup_key(part) for part in item[0])):
        out.append(
            [
                key[0],
                key[1],
                key[2],
                key[3],
                key[4],
                summary["rows"],
                round(summary["hours"], 4),
                round(summary["total"], 2),
            ]
        )
    return out


def remove_sheet_if_present(workbook, sheet_name: str) -> None:
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])


def populate_balance_sheet(workbook, sheet_name: str, data_rows: list[list[Any]], hidden: bool = False) -> None:
    remove_sheet_if_present(workbook, sheet_name)
    ws = workbook.create_sheet(sheet_name)
    ws.sheet_state = "hidden" if hidden else "visible"

    for col_index, header in enumerate(BALANCE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_index)
        cell.value = header

    balance_rows = build_balance_rows(data_rows, sheet_name)
    if not balance_rows:
        balance_rows = [["", "", "", "", "", 0, 0, 0]]

    for row_index, values in enumerate(balance_rows, start=2):
        for col_index, value in enumerate(values, start=1):
            set_cell_value(ws.cell(row=row_index, column=col_index), value)

    for column, width in {
        "A": 24,
        "B": 28,
        "C": 34,
        "D": 42,
        "E": 22,
        "F": 10,
        "G": 18,
        "H": 16,
    }.items():
        ws.column_dimensions[column].width = width

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=7).number_format = "0.0000"
        ws.cell(row=row, column=8).number_format = "0.00"


def populate_balance_sheets(workbook, data_rows: list[list[Any]]) -> None:
    populate_balance_sheet(workbook, PAGOS_PERIODOS_BALANCE_SHEET_NAME, data_rows)
    populate_balance_sheet(workbook, TICKETS_BALANCE_SHEET_NAME, data_rows)
    populate_balance_sheet(workbook, BANK_PERIODICS_BALANCE_SHEET_NAME, data_rows, hidden=True)


def finalize_workbook_sheets(workbook, generated_sheet_names: list[str]) -> None:
    for sheet_name in list(workbook.sheetnames):
        if sheet_name == WORK98_TEMPLATE_SHEET_NAME or (
            sheet_name not in generated_sheet_names
            and sheet_name != DATA_TEMPLATE_SHEET_NAME
            and sheet_name not in BALANCE_SHEET_NAMES
        ):
            if sheet_name in generated_sheet_names:
                continue
            workbook.remove(workbook[sheet_name])


def strip_external_formulas(target_sheet, value_sheet) -> None:
    for row in target_sheet.iter_rows():
        for cell in row:
            if cell.data_type == "f" and isinstance(cell.value, str) and "[" in cell.value:
                set_cell_value(target_sheet[cell.coordinate], value_sheet[cell.coordinate].value)


def build_work98_workbook_bytes(
    template_content: bytes,
    jobs: list[dict[str, Any]],
    contractors: dict[str, dict[str, str]],
    buildings: dict[str, dict[str, str]],
) -> bytes:
    workbook = load_workbook(BytesIO(template_content))
    values_workbook = load_workbook(BytesIO(template_content), data_only=True)
    if WORK98_TEMPLATE_SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(f'Template sheet "{WORK98_TEMPLATE_SHEET_NAME}" not found')

    template_sheet = workbook[WORK98_TEMPLATE_SHEET_NAME]
    values_template_sheet = values_workbook[WORK98_TEMPLATE_SHEET_NAME]
    data_sheet = ensure_data_sheet(workbook)
    generated_sheet_names: list[str] = []
    generated_data_rows: list[list[Any]] = []
    work_index = 1

    for plan_jobs in ordered_plan_groups(jobs, contractors):
        plan_render_rows = build_plan_render_rows(plan_jobs, contractors)
        for plan_chunk in chunk_render_rows(plan_render_rows, WORK98_LINE_CAPACITY):
            plan_chunk_jobs = [job for job in plan_chunk if job is not None]
            sheet = workbook.copy_worksheet(template_sheet)
            sheet.title = f"{WORK98_OUTPUT_PREFIX}{work_index}"
            strip_external_formulas(sheet, values_template_sheet)
            populate_work98_header(sheet, plan_chunk_jobs, buildings)
            populate_work98_rows(sheet, plan_chunk, contractors, buildings)
            for job in plan_chunk_jobs:
                contractor = find_contractor_lookup(job, contractors)
                building = find_building_lookup(job, buildings)
                generated_data_rows.append(build_data_row(sheet.title, job, contractor, building))
            generated_sheet_names.append(sheet.title)
            work_index += 1

    if not generated_sheet_names:
        sheet = workbook.copy_worksheet(template_sheet)
        sheet.title = f"{WORK98_OUTPUT_PREFIX}1"
        strip_external_formulas(sheet, values_template_sheet)
        clear_work98_line_rows(sheet)
        generated_sheet_names.append(sheet.title)

    populate_data_sheet(data_sheet, generated_data_rows)
    finalize_workbook_sheets(workbook, generated_sheet_names)
    populate_balance_sheets(workbook, generated_data_rows)
    workbook._sheets = (
        [sheet for sheet in workbook._sheets if sheet.title not in (*BALANCE_SHEET_NAMES, DATA_TEMPLATE_SHEET_NAME)]
        + [workbook[PAGOS_PERIODOS_BALANCE_SHEET_NAME], workbook[TICKETS_BALANCE_SHEET_NAME]]
        + [data_sheet, workbook[BANK_PERIODICS_BALANCE_SHEET_NAME]]
    )
    workbook.defined_names.clear()
    workbook._external_links = []
    workbook.active = 0

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def generate_work98_workbook(
    token: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
) -> tuple[str, bytes]:
    jobs, contractors, buildings = load_work98_inputs()
    jobs = filter_jobs_by_date_range(jobs, start_date=start_date, end_date=end_date)
    template_content = work98_template_bytes(token)
    workbook_content = build_work98_workbook_bytes(template_content, jobs, contractors, buildings)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"work_export_{timestamp}.xlsx", workbook_content


def generate_work98_raw_export(
    start_date: date | None = None,
    end_date: date | None = None,
) -> tuple[str, bytes]:
    jobs, _, _ = load_work98_inputs()
    jobs = filter_jobs_by_date_range(jobs, start_date=start_date, end_date=end_date)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"work_export_raw_{timestamp}.csv", build_raw_jobs_csv_bytes(jobs)


def fetch_schedule_validation_rows() -> list[dict[str, Any]]:
    sql = """
        SELECT
            s.id AS schedule_instance_id,
            BIN_TO_UUID(s.route_id) AS schedule_route_id,
            BIN_TO_UUID(s.group_id) AS schedule_group_id,
            BIN_TO_UUID(s.plan_id) AS schedule_plan_id,
            BIN_TO_UUID(s.plan_user_id) AS schedule_plan_user_id,
            BIN_TO_UUID(s.worker_user_id) AS schedule_worker_user_id,
            BIN_TO_UUID(s.supervisor_id) AS schedule_supervisor_id,
            s.worker_name AS schedule_worker_name,
            s.route_name AS schedule_route_name,
            s.route_description AS schedule_route_description,
            s.plan_description AS schedule_plan_description,
            s.plan_type_name AS schedule_plan_type_name,
            s.building_label AS schedule_building_label,
            s.supervisor_name AS schedule_supervisor_name,
            s.timezone_name AS schedule_timezone_name,
            s.scheduled_start_at AS schedule_start_at,
            s.scheduled_end_at AS schedule_end_at,
            s.schedule_status AS schedule_status,
            s.refreshed_at AS schedule_refreshed_at,
            s.status_changed_at AS schedule_status_changed_at,
            rav.current AS route_assignment_current,
            rav.route_name AS route_assignment_route_name,
            rav.route_description AS route_assignment_route_description,
            rav.pagos_periodos_id AS route_assignment_pagos_periodos_id,
            rav.type_of_payment AS route_assignment_type_of_payment,
            rav.hourly_rate AS route_assignment_hourly_rate,
            rav.schedule_label AS route_assignment_schedule_label,
            rav.scheduled_hours AS route_assignment_scheduled_hours,
            BIN_TO_UUID(rav.id) AS route_assignment_version_id,
            pu.active AS plan_user_active,
            pu.has_schedule AS plan_user_has_schedule,
            pu.assigned_at AS plan_user_assigned_at,
            COALESCE(p.first_name, '') AS worker_first_name,
            COALESCE(p.last_name, '') AS worker_last_name,
            COALESCE(u.username, '') AS worker_username,
            pl.status AS plan_status,
            pl.start_date AS plan_start_date,
            pl.end_date AS plan_end_date,
            pl.description AS plan_description,
            pl.po_code AS plan_po_code,
            pl.budget AS plan_budget,
            pt.name AS plan_type_name,
            BIN_TO_UUID(b.id) AS building_id,
            COALESCE(b.public_building_id, '') AS building_number,
            COALESCE(b.public_expense_id, '') AS building_expense_id,
            b.address AS building_address,
            COALESCE(b.pagos_periodos_id, '') AS building_pagos_periodos_id,
            COALESCE(c.city_name, '') AS building_city,
            COALESCE(pr.province_name, '') AS building_province,
            COALESCE(co.country_name, '') AS building_country
        FROM scheduled_shift_instances s
        LEFT JOIN route_assignment_versions rav
            ON rav.route_id = s.route_id
           AND rav.plan_user_id = s.plan_user_id
           AND rav.current = 1
        LEFT JOIN plan_users pu
            ON pu.id = s.plan_user_id
        LEFT JOIN users u
            ON u.id = pu.user_id
        LEFT JOIN persons p
            ON p.id = u.person_id
        LEFT JOIN plans pl
            ON pl.id = s.plan_id
        LEFT JOIN plan_types pt
            ON pt.id = pl.type_id
        LEFT JOIN buildings b
            ON b.id = pl.building_id
        LEFT JOIN cities c
            ON c.id = b.city_id
        LEFT JOIN provinces pr
            ON pr.id = c.province_id
        LEFT JOIN countries co
            ON co.id = pr.country_id
        ORDER BY s.scheduled_start_at DESC, s.worker_name ASC, s.route_name ASC
    """
    with mysql_connection() as connection:
        with connection.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute(sql)
            return list(cursor.fetchall())


def generate_work98_schedule_validation_export() -> tuple[str, bytes]:
    rows = fetch_schedule_validation_rows()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"work_export_schedule_validation_{timestamp}.csv", build_dict_rows_csv_bytes(rows)
