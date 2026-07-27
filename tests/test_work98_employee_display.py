import unittest
from datetime import datetime

from openpyxl import Workbook

from utils.work98_generator import (
    EMPLOYEE_DATA_NAME_COLUMN,
    ORANGE_FILL,
    REVIEW_NO_FILL,
    REVIEW_YES_FILL,
    ATTENDED_COLUMN,
    LATE_COLUMN,
    REVIEW_COLUMN,
    SCHEDULE_END_COLUMN,
    SCHEDULE_HOURS_COLUMN,
    SCHEDULE_HOURS_DIFF_COLUMN,
    build_data_row,
    employee_display_for_job,
    populate_work98_rows,
    resolved_work_sheet_end_time,
)


class Work98EmployeeDisplayTests(unittest.TestCase):
    def test_combines_cnetbms_name_and_pagos_periodos_id(self):
        job = {
            "worker_first_name": "Abigail",
            "worker_last_name": "Lucero",
            "assigned_user_pagos_periodos_id": "Aaron Guzman",
        }

        self.assertEqual(
            "Abigail Lucero (Aaron Guzman)",
            employee_display_for_job(job, {}),
        )

    def test_uses_cnetbms_name_when_pagos_periodos_id_is_empty(self):
        job = {
            "worker_first_name": "Abigail",
            "worker_last_name": "Lucero",
            "assigned_user_pagos_periodos_id": "",
        }

        self.assertEqual("Abigail Lucero", employee_display_for_job(job, {}))

    def test_uses_pagos_periodos_id_when_cnetbms_name_is_empty(self):
        job = {
            "worker_first_name": "",
            "worker_last_name": "",
            "worker_username": "",
            "assigned_user_pagos_periodos_id": "Aaron Guzman",
        }

        self.assertEqual("Aaron Guzman", employee_display_for_job(job, {}))

    def test_populates_combined_name_only_in_work_sheet(self):
        job = {
            "worker_first_name": "Abigail",
            "worker_last_name": "Lucero",
            "worker_username": "alucero",
            "assigned_user_pagos_periodos_id": "Aaron Guzman",
        }
        contractor = {
            "name": "Aaron Guzman",
            "hourly_rate": "17.75",
            "type_of_payment": "",
            "employee_sub_grouping": "Aaron Guzman",
            "category": "Labor T4A",
        }
        contractors = {"aaron guzman": contractor}
        worksheet = Workbook().active

        populate_work98_rows(worksheet, [job], contractors, {})
        data_row = build_data_row("Work1", job, contractor, None)

        self.assertEqual("Abigail Lucero (Aaron Guzman)", worksheet["C6"].value)
        self.assertEqual("Abigail Lucero", worksheet.cell(6, EMPLOYEE_DATA_NAME_COLUMN).value)
        self.assertEqual("Aaron Guzman", data_row[5])

    def test_out_uses_real_time_when_no_truncation_exists(self):
        job = {
            "job_needs_review": 0,
            "job_end_time": datetime(2026, 7, 24, 2, 20),
        }

        result = resolved_work_sheet_end_time(job, None)

        self.assertEqual((2026, 7, 23, 22, 20), self._date_time_parts(result))

    def test_out_uses_trimmed_time_when_review_is_no(self):
        job = {
            "job_needs_review": 0,
            "job_trim_end": datetime(2026, 7, 24, 2, 30),
            "job_end_time": datetime(2026, 7, 24, 2, 40),
        }

        result = resolved_work_sheet_end_time(job, None)

        self.assertEqual((2026, 7, 23, 22, 30), self._date_time_parts(result))

    def test_out_uses_manual_time_only_when_review_is_no(self):
        job = {
            "job_needs_review": 0,
            "job_manual_end": datetime(2026, 7, 24, 2, 25),
            "job_trim_end": datetime(2026, 7, 24, 2, 30),
            "job_end_time": datetime(2026, 7, 24, 2, 40),
        }

        result = resolved_work_sheet_end_time(job, None)

        self.assertEqual((2026, 7, 23, 22, 25), self._date_time_parts(result))

    def test_out_uses_real_time_when_review_is_yes(self):
        job = {
            "job_needs_review": 1,
            "job_trim_end": datetime(2026, 7, 24, 2, 30),
            "job_end_time": datetime(2026, 7, 24, 2, 45),
        }

        result = resolved_work_sheet_end_time(job, None)

        self.assertEqual((2026, 7, 23, 22, 45), self._date_time_parts(result))

    def test_manual_time_does_not_override_real_time_when_review_is_yes(self):
        job = {
            "job_needs_review": 1,
            "job_manual_end": datetime(2026, 7, 24, 2, 25),
            "job_trim_end": datetime(2026, 7, 24, 2, 30),
            "job_end_time": datetime(2026, 7, 24, 2, 45),
        }

        result = resolved_work_sheet_end_time(job, None)

        self.assertEqual((2026, 7, 23, 22, 45), self._date_time_parts(result))

    def test_work_sheet_out_cell_uses_real_time_when_review_is_yes(self):
        job = {
            "job_needs_review": 1,
            "job_manual_end": datetime(2026, 7, 24, 2, 25),
            "job_trim_end": datetime(2026, 7, 24, 2, 30),
            "job_end_time": datetime(2026, 7, 24, 2, 45),
        }
        worksheet = Workbook().active

        populate_work98_rows(worksheet, [job], {}, {})

        self.assertEqual(
            (2026, 7, 23, 22, 45),
            self._date_time_parts(worksheet["E6"].value),
        )

    def test_work_sheet_colors_scheduled_out_orange_and_review_yes_green(self):
        job = {
            "job_needs_review": 1,
            "job_scheduled_end": datetime(2026, 7, 24, 2, 30),
        }
        worksheet = Workbook().active

        populate_work98_rows(worksheet, [job], {}, {})

        self.assertEqual(
            ORANGE_FILL.fgColor.rgb,
            worksheet.cell(6, SCHEDULE_END_COLUMN).fill.fgColor.rgb,
        )
        self.assertEqual(
            REVIEW_YES_FILL.fgColor.rgb,
            worksheet.cell(6, REVIEW_COLUMN).fill.fgColor.rgb,
        )

    def test_work_sheet_colors_review_no_red(self):
        worksheet = Workbook().active

        populate_work98_rows(worksheet, [{"job_needs_review": 0}], {}, {})

        self.assertEqual(
            REVIEW_NO_FILL.fgColor.rgb,
            worksheet.cell(6, REVIEW_COLUMN).fill.fgColor.rgb,
        )

    def test_work_sheet_leaves_empty_review_without_a_solid_fill(self):
        worksheet = Workbook().active

        populate_work98_rows(worksheet, [{"job_needs_review": None}], {}, {})

        self.assertNotEqual(
            "solid",
            worksheet.cell(6, REVIEW_COLUMN).fill.fill_type,
        )

    def test_schedule_validation_columns_contain_values_not_uncalculated_formulas(self):
        job = {
            "job_needs_review": 0,
            "job_start_time": datetime(2026, 7, 21, 10, 23, 44),
            "job_end_time": datetime(2026, 7, 21, 19, 3, 46),
            "job_trim_start": datetime(2026, 7, 21, 11, 0),
            "job_trim_end": datetime(2026, 7, 21, 19, 0),
            "job_scheduled_start": datetime(2026, 7, 21, 11, 0),
            "job_scheduled_end": datetime(2026, 7, 21, 19, 0),
        }
        worksheet = Workbook().active

        populate_work98_rows(worksheet, [job], {}, {})

        self.assertEqual(7.5, worksheet.cell(6, SCHEDULE_HOURS_COLUMN).value)
        self.assertEqual(0.0, worksheet.cell(6, SCHEDULE_HOURS_DIFF_COLUMN).value)
        self.assertEqual("Yes", worksheet.cell(6, ATTENDED_COLUMN).value)
        self.assertEqual("No", worksheet.cell(6, LATE_COLUMN).value)
        for column in (
            SCHEDULE_HOURS_COLUMN,
            SCHEDULE_HOURS_DIFF_COLUMN,
            ATTENDED_COLUMN,
            LATE_COLUMN,
        ):
            self.assertNotEqual("f", worksheet.cell(6, column).data_type)

    @staticmethod
    def _date_time_parts(value):
        return (value.year, value.month, value.day, value.hour, value.minute)


if __name__ == "__main__":
    unittest.main()
